"""
Excel and Google Sheets Export Utility

This module provides functionality to export candidate data to Excel files
and Google Sheets for organized storage and analysis.
"""

import json
import logging
import os
from datetime import datetime
from typing import Any, Dict, List, Optional, Union

import pandas as pd

logger = logging.getLogger(__name__)

# Try to import Google Sheets dependencies
try:
    import gspread
    from google.oauth2.service_account import Credentials
    GOOGLE_SHEETS_AVAILABLE = True
except ImportError:
    GOOGLE_SHEETS_AVAILABLE = False
    logger.warning("Google Sheets dependencies not available. Install with: pip install gspread google-auth")


class ExportManager:
    """
    Manages exporting candidate data to Excel and Google Sheets.
    
    Features:
    - Export to Excel (.xlsx) files
    - Export to Google Sheets (with proper authentication)
    - Organized data structure with multiple sheets
    - Automatic formatting and styling
    - Batch processing support
    """
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """
        Initialize the export manager.
        
        Args:
            config: Configuration dictionary
        """
        self.config = config or {}
        self.google_sheets_client = None
        
        # Initialize Google Sheets if available
        if GOOGLE_SHEETS_AVAILABLE:
            self._initialize_google_sheets()
    
    def _initialize_google_sheets(self) -> None:
        """Initialize Google Sheets client with authentication."""
        try:
            # Check for service account credentials
            creds_file = self.config.get('GOOGLE_SERVICE_ACCOUNT_FILE', 'service_account.json')
            
            if os.path.exists(creds_file):
                scope = [
                    'https://spreadsheets.google.com/feeds',
                    'https://www.googleapis.com/auth/drive'
                ]
                
                creds = Credentials.from_service_account_file(creds_file, scopes=scope)
                self.google_sheets_client = gspread.authorize(creds)
                logger.info("Google Sheets client initialized successfully")
            else:
                logger.warning(f"Google service account file not found: {creds_file}")
                
        except Exception as e:
            logger.error(f"Failed to initialize Google Sheets client: {str(e)}")
    
    def export_to_excel(
        self,
        candidates: List[Dict[str, Any]],
        output_file: str,
        include_analytics: bool = True,
        include_messages: bool = True
    ) -> bool:
        """
        Export candidate data to Excel file with multiple organized sheets.
        
        Args:
            candidates: List of candidate dictionaries
            output_file: Path to output Excel file
            include_analytics: Whether to include analytics sheet
            include_messages: Whether to include generated messages
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Create Excel writer
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                
                # 1. Main Candidates Sheet
                candidates_df = self._prepare_candidates_dataframe(candidates)
                candidates_df.to_excel(writer, sheet_name='Candidates', index=False)
                
                # 2. Contact Information Sheet
                contact_df = self._prepare_contact_dataframe(candidates)
                contact_df.to_excel(writer, sheet_name='Contact_Info', index=False)
                
                # 3. Experience & Education Sheet  
                experience_df = self._prepare_experience_dataframe(candidates)
                experience_df.to_excel(writer, sheet_name='Experience_Education', index=False)
                
                # 4. Skills & Scoring Sheet
                skills_df = self._prepare_skills_dataframe(candidates)
                skills_df.to_excel(writer, sheet_name='Skills_Scoring', index=False)
                
                # 5. Multi-Source Data Sheet
                multi_source_df = self._prepare_multi_source_dataframe(candidates)
                multi_source_df.to_excel(writer, sheet_name='Multi_Source_Data', index=False)
                
                # 6. Generated Messages Sheet (if requested)
                if include_messages:
                    messages_df = self._prepare_messages_dataframe(candidates)
                    messages_df.to_excel(writer, sheet_name='Generated_Messages', index=False)
                
                # 7. Analytics Sheet (if requested)
                if include_analytics:
                    analytics_df = self._prepare_analytics_dataframe(candidates)
                    analytics_df.to_excel(writer, sheet_name='Analytics', index=False)
                
                # 8. Summary Sheet
                summary_df = self._prepare_summary_dataframe(candidates)
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Apply formatting
            self._format_excel_file(output_file)
            
            logger.info(f"Successfully exported {len(candidates)} candidates to {output_file}")
            return True
            
        except Exception as e:
            logger.error(f"Failed to export to Excel: {str(e)}")
            return False
    
    def export_to_google_sheets(
        self,
        candidates: List[Dict[str, Any]],
        spreadsheet_name: str,
        share_with_email: Optional[str] = None
    ) -> Optional[str]:
        """
        Export candidate data to Google Sheets.
        
        Args:
            candidates: List of candidate dictionaries
            spreadsheet_name: Name for the Google Spreadsheet
            share_with_email: Email to share the spreadsheet with
            
        Returns:
            Spreadsheet URL if successful, None otherwise
        """
        if not self.google_sheets_client:
            logger.error("Google Sheets client not initialized")
            return None
        
        try:
            # Create new spreadsheet
            spreadsheet = self.google_sheets_client.create(spreadsheet_name)
            
            # Share with email if provided
            if share_with_email:
                spreadsheet.share(share_with_email, perm_type='user', role='writer')
            
            # Get all sheets we need to create
            sheet_data = self._prepare_all_sheets_data(candidates)
            
            # Create and populate sheets
            for i, (sheet_name, df) in enumerate(sheet_data.items()):
                if i == 0:
                    # Use the default sheet for the first one
                    worksheet = spreadsheet.sheet1
                    worksheet.update_title(sheet_name)
                else:
                    # Create new sheets for the rest
                    worksheet = spreadsheet.add_worksheet(title=sheet_name, rows=len(df)+10, cols=len(df.columns)+5)
                
                # Convert DataFrame to list of lists for Google Sheets
                data = [df.columns.tolist()] + df.values.tolist()
                
                # Update the worksheet
                worksheet.update('A1', data)
                
                # Format headers
                self._format_google_sheet_headers(worksheet, len(df.columns))
            
            logger.info(f"Successfully exported {len(candidates)} candidates to Google Sheets: {spreadsheet.url}")
            return spreadsheet.url
            
        except Exception as e:
            logger.error(f"Failed to export to Google Sheets: {str(e)}")
            return None
    
    def _prepare_candidates_dataframe(self, candidates: List[Dict[str, Any]]) -> pd.DataFrame:
        """Prepare main candidates dataframe."""
        data = []
        for candidate in candidates:
            row = {
                'Name': candidate.get('name', ''),
                'Headline': candidate.get('headline', ''),
                'Location': candidate.get('location', ''),
                'Company': self._extract_current_company(candidate),
                'Title': self._extract_current_title(candidate),
                'LinkedIn_URL': candidate.get('linkedin_url') or candidate.get('profile_url', ''),
                'Fit_Score': candidate.get('fit_score', 0),
                'Confidence': candidate.get('confidence', ''),
                'Status': candidate.get('status', 'New'),
                'Date_Added': candidate.get('date_added', datetime.now().strftime('%Y-%m-%d')),
                'Notes': candidate.get('notes', '')
            }
            data.append(row)
        
        return pd.DataFrame(data)
    
    def _prepare_contact_dataframe(self, candidates: List[Dict[str, Any]]) -> pd.DataFrame:
        """Prepare contact information dataframe."""
        data = []
        for candidate in candidates:
            row = {
                'Name': candidate.get('name', ''),
                'Email': candidate.get('email', ''),
                'Phone': candidate.get('phone', ''),
                'LinkedIn_URL': candidate.get('linkedin_url') or candidate.get('profile_url', ''),
                'Personal_Website': candidate.get('personal_website', {}).get('url', ''),
                'GitHub_Username': candidate.get('github_profile', {}).get('username', ''),
                'Twitter_Username': candidate.get('twitter_profile', {}).get('username', ''),
                'Location': candidate.get('location', ''),
                'Time_Zone': candidate.get('timezone', ''),
                'Preferred_Contact': candidate.get('preferred_contact_method', 'LinkedIn')
            }
            data.append(row)
        
        return pd.DataFrame(data)
    
    def _prepare_experience_dataframe(self, candidates: List[Dict[str, Any]]) -> pd.DataFrame:
        """Prepare experience and education dataframe."""
        data = []
        for candidate in candidates:
            # Experience
            experience = candidate.get('experience', [])
            experience_summary = self._summarize_experience(experience)
            
            # Education
            education = candidate.get('education', [])
            education_summary = self._summarize_education(education)
            
            row = {
                'Name': candidate.get('name', ''),
                'Current_Company': self._extract_current_company(candidate),
                'Current_Title': self._extract_current_title(candidate),
                'Years_Experience': self._calculate_years_experience(experience),
                'Previous_Companies': self._extract_previous_companies(experience),
                'Experience_Summary': experience_summary,
                'Education_Level': self._extract_education_level(education),
                'Schools': self._extract_schools(education),
                'Degrees': self._extract_degrees(education),
                'Education_Summary': education_summary,
                'Certifications': candidate.get('certifications', [])
            }
            data.append(row)
        
        return pd.DataFrame(data)
    
    def _prepare_skills_dataframe(self, candidates: List[Dict[str, Any]]) -> pd.DataFrame:
        """Prepare skills and scoring dataframe."""
        data = []
        for candidate in candidates:
            skills = candidate.get('skills', [])
            
            row = {
                'Name': candidate.get('name', ''),
                'Fit_Score': candidate.get('fit_score', 0),
                'Technical_Skills': ', '.join(skills[:10]) if skills else '',
                'All_Skills': ', '.join(skills) if skills else '',
                'Skills_Count': len(skills),
                'Matching_Keywords': ', '.join(candidate.get('matching_keywords', [])),
                'Relevance_Score': candidate.get('relevance_score', 0),
                'Experience_Score': candidate.get('experience_score', 0),
                'Education_Score': candidate.get('education_score', 0),
                'Overall_Rating': candidate.get('overall_rating', 'Unrated'),
                'Strengths': ', '.join(candidate.get('strengths', [])),
                'Potential_Concerns': ', '.join(candidate.get('concerns', []))
            }
            data.append(row)
        
        return pd.DataFrame(data)
    
    def _prepare_multi_source_dataframe(self, candidates: List[Dict[str, Any]]) -> pd.DataFrame:
        """Prepare multi-source data dataframe."""
        data = []
        for candidate in candidates:
            github = candidate.get('github_profile', {})
            twitter = candidate.get('twitter_profile', {})
            website = candidate.get('personal_website', {})
            
            row = {
                'Name': candidate.get('name', ''),
                'Has_GitHub': bool(github),
                'GitHub_Username': github.get('username', ''),
                'GitHub_Repos': github.get('public_repos', 0),
                'GitHub_Stars': sum(repo.get('stars', 0) for repo in github.get('notable_repos', [])),
                'GitHub_Languages': ', '.join(github.get('top_languages', [])),
                'Has_Twitter': bool(twitter),
                'Twitter_Username': twitter.get('username', ''),
                'Twitter_Followers': twitter.get('followers', 0),
                'Twitter_Bio': twitter.get('bio', ''),
                'Has_Website': bool(website),
                'Website_URL': website.get('url', ''),
                'Has_Blog': website.get('has_blog', False),
                'Has_Portfolio': website.get('has_portfolio', False),
                'Content_Topics': ', '.join(website.get('content_topics', [])),
                'Social_Media_Score': candidate.get('social_media_score', 0)
            }
            data.append(row)
        
        return pd.DataFrame(data)
    
    def _prepare_messages_dataframe(self, candidates: List[Dict[str, Any]]) -> pd.DataFrame:
        """Prepare generated messages dataframe."""
        data = []
        for candidate in candidates:
            # Check both possible message fields
            messages = candidate.get('generated_messages', [])
            outreach_message = candidate.get('outreach_message', '')
            
            if outreach_message:
                # Handle direct outreach_message field
                row = {
                    'Name': candidate.get('name', ''),
                    'Message_Type': 'LinkedIn Outreach',
                    'Message_Content': outreach_message,
                    'Generation_Method': candidate.get('generation_method', 'Template'),
                    'Personalization_Score': candidate.get('personalization_score', 3),
                    'Confidence': candidate.get('confidence', 'Medium'),
                    'Generated_Date': candidate.get('scoring_timestamp', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                    'Template_Used': candidate.get('template_used', 'professional_outreach'),
                    'Character_Count': len(outreach_message)
                }
                data.append(row)
            elif messages:
                # Handle structured messages array
                for message in messages:
                    row = {
                        'Name': candidate.get('name', ''),
                        'Message_Type': message.get('message_type', ''),
                        'Message_Content': message.get('message', ''),
                        'Generation_Method': message.get('generation_method', ''),
                        'Personalization_Score': message.get('personalization_score', 0),
                        'Confidence': message.get('confidence', ''),
                        'Generated_Date': message.get('generation_timestamp', ''),
                        'Template_Used': message.get('template_used', ''),
                        'Character_Count': len(message.get('message', ''))
                    }
                    data.append(row)
            else:
                # Create empty row if no messages
                row = {
                    'Name': candidate.get('name', ''),
                    'Message_Type': 'No Message Generated',
                    'Message_Content': 'No outreach message available',
                    'Generation_Method': 'None',
                    'Personalization_Score': 0,
                    'Confidence': 'N/A',
                    'Generated_Date': '',
                    'Template_Used': 'None',
                    'Character_Count': 0
                }
                data.append(row)
        
        return pd.DataFrame(data)
    
    def _prepare_analytics_dataframe(self, candidates: List[Dict[str, Any]]) -> pd.DataFrame:
        """Prepare analytics dataframe."""
        if not candidates:
            return pd.DataFrame()
        
        # Calculate analytics
        total_candidates = len(candidates)
        avg_fit_score = sum(c.get('fit_score', 0) for c in candidates) / total_candidates if total_candidates > 0 else 0
        
        # Score distribution
        high_fit = len([c for c in candidates if c.get('fit_score', 0) >= 8])
        medium_fit = len([c for c in candidates if 5 <= c.get('fit_score', 0) < 8])
        low_fit = len([c for c in candidates if c.get('fit_score', 0) < 5])
        
        # Location distribution
        locations = {}
        for candidate in candidates:
            location = candidate.get('location', 'Unknown')
            locations[location] = locations.get(location, 0) + 1
        
        # Experience distribution
        experience_levels = {}
        for candidate in candidates:
            level = self._categorize_experience_level(candidate)
            experience_levels[level] = experience_levels.get(level, 0) + 1
        
        # Create analytics data
        analytics_data = [
            {'Metric': 'Total Candidates', 'Value': total_candidates, 'Percentage': '100%'},
            {'Metric': 'Average Fit Score', 'Value': round(avg_fit_score, 2), 'Percentage': ''},
            {'Metric': 'High Fit Candidates (8+)', 'Value': high_fit, 'Percentage': f'{high_fit/total_candidates*100:.1f}%' if total_candidates > 0 else '0%'},
            {'Metric': 'Medium Fit Candidates (5-7)', 'Value': medium_fit, 'Percentage': f'{medium_fit/total_candidates*100:.1f}%' if total_candidates > 0 else '0%'},
            {'Metric': 'Low Fit Candidates (<5)', 'Value': low_fit, 'Percentage': f'{low_fit/total_candidates*100:.1f}%' if total_candidates > 0 else '0%'},
        ]
        
        # Add location data
        for location, count in sorted(locations.items(), key=lambda x: x[1], reverse=True)[:5]:
            analytics_data.append({
                'Metric': f'Location: {location}',
                'Value': count,
                'Percentage': f'{count/total_candidates*100:.1f}%' if total_candidates > 0 else '0%'
            })
        
        # Add experience level data
        for level, count in sorted(experience_levels.items(), key=lambda x: x[1], reverse=True):
            analytics_data.append({
                'Metric': f'Experience: {level}',
                'Value': count,
                'Percentage': f'{count/total_candidates*100:.1f}%' if total_candidates > 0 else '0%'
            })
        
        return pd.DataFrame(analytics_data)
    
    def _prepare_summary_dataframe(self, candidates: List[Dict[str, Any]]) -> pd.DataFrame:
        """Prepare summary dataframe."""
        if not candidates:
            return pd.DataFrame()
        
        # Create summary data
        summary_data = []
        
        # Export metadata
        summary_data.append({'Field': 'Export Date', 'Value': datetime.now().strftime('%Y-%m-%d %H:%M:%S')})
        summary_data.append({'Field': 'Total Candidates', 'Value': len(candidates)})
        summary_data.append({'Field': 'Search Query', 'Value': candidates[0].get('search_query', 'N/A') if candidates else 'N/A'})
        summary_data.append({'Field': 'Job Title', 'Value': candidates[0].get('job_title', 'N/A') if candidates else 'N/A'})
        
        # Top performing candidates
        top_candidates = sorted(candidates, key=lambda x: x.get('fit_score', 0), reverse=True)[:5]
        for i, candidate in enumerate(top_candidates, 1):
            summary_data.append({
                'Field': f'Top Candidate #{i}',
                'Value': f"{candidate.get('name', 'Unknown')} (Score: {candidate.get('fit_score', 0)})"
            })
        
        return pd.DataFrame(summary_data)
    
    def _prepare_all_sheets_data(self, candidates: List[Dict[str, Any]]) -> Dict[str, pd.DataFrame]:
        """Prepare all sheets data for Google Sheets export."""
        return {
            'Candidates': self._prepare_candidates_dataframe(candidates),
            'Contact_Info': self._prepare_contact_dataframe(candidates),
            'Experience_Education': self._prepare_experience_dataframe(candidates),
            'Skills_Scoring': self._prepare_skills_dataframe(candidates),
            'Multi_Source_Data': self._prepare_multi_source_dataframe(candidates),
            'Generated_Messages': self._prepare_messages_dataframe(candidates),
            'Analytics': self._prepare_analytics_dataframe(candidates),
            'Summary': self._prepare_summary_dataframe(candidates)
        }
    
    def _format_excel_file(self, file_path: str) -> None:
        """Apply formatting to Excel file."""
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            workbook = load_workbook(file_path)
            
            # Format each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Header formatting
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                # Apply header formatting
                for cell in sheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Auto-adjust column widths
                for column in sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    sheet.column_dimensions[column_letter].width = adjusted_width
            
            workbook.save(file_path)
            logger.info(f"Applied formatting to {file_path}")
            
        except Exception as e:
            logger.warning(f"Failed to apply Excel formatting: {str(e)}")
    
    def _format_google_sheet_headers(self, worksheet, num_columns: int) -> None:
        """Format headers in Google Sheet."""
        try:
            # Format header row
            worksheet.format('A1:' + chr(ord('A') + num_columns - 1) + '1', {
                'backgroundColor': {'red': 0.2, 'green': 0.38, 'blue': 0.57},
                'textFormat': {'foregroundColor': {'red': 1, 'green': 1, 'blue': 1}, 'bold': True},
                'horizontalAlignment': 'CENTER'
            })
            
        except Exception as e:
            logger.warning(f"Failed to format Google Sheet headers: {str(e)}")
    
    # Helper methods
    def _extract_current_company(self, candidate: Dict[str, Any]) -> str:
        """Extract current company from candidate data."""
        headline = candidate.get('headline', '')
        
        # Try different separators in order of priority
        separators = [' at ', ' @ ']
        for sep in separators:
            if sep in headline:
                company_part = headline.split(sep)[1]
                # Clean up company name by removing additional info after • | - ( )
                # Split by • first to get just the company name before technical details
                if '•' in company_part:
                    company_clean = company_part.split('•')[0].strip()
                elif '|' in company_part:
                    company_clean = company_part.split('|')[0].strip()
                elif '-' in company_part:
                    company_clean = company_part.split('-')[0].strip()
                elif '(' in company_part:
                    company_clean = company_part.split('(')[0].strip()
                else:
                    company_clean = company_part.strip()
                return company_clean
        
        # If no separators found, try to extract company from experience
        experience = candidate.get('experience', [])
        if experience and isinstance(experience[0], dict):
            return experience[0].get('company', '')
        
        return ''
    
    def _extract_current_title(self, candidate: Dict[str, Any]) -> str:
        """Extract current title from candidate data."""
        headline = candidate.get('headline', '')
        separators = [' at ', ' | ', ' - ', ' @ ']
        for sep in separators:
            if sep in headline:
                return headline.split(sep)[0].strip()
        return headline.strip()
    
    def _summarize_experience(self, experience: List[Any]) -> str:
        """Create experience summary."""
        if not experience:
            return ''
        
        summaries = []
        for exp in experience[:3]:
            if isinstance(exp, dict):
                title = exp.get('title', '')
                company = exp.get('company', '')
                if title and company:
                    summaries.append(f"{title} at {company}")
                elif title:
                    summaries.append(title)
        
        return '; '.join(summaries)
    
    def _summarize_education(self, education: List[Any]) -> str:
        """Create education summary."""
        if not education:
            return ''
        
        summaries = []
        for edu in education[:2]:
            if isinstance(edu, dict):
                degree = edu.get('degree', '')
                school = edu.get('school', '')
                if degree and school:
                    summaries.append(f"{degree} from {school}")
                elif school:
                    summaries.append(school)
        
        return '; '.join(summaries)
    
    def _calculate_years_experience(self, experience: List[Any]) -> int:
        """Calculate years of experience."""
        # Simplified calculation - in real implementation, parse dates
        return len(experience) * 2 if experience else 0
    
    def _extract_previous_companies(self, experience: List[Any]) -> str:
        """Extract previous companies."""
        companies = []
        for exp in experience[1:4]:  # Skip current (first) and take next 3
            if isinstance(exp, dict):
                company = exp.get('company', '')
                if company:
                    companies.append(company)
        return ', '.join(companies)
    
    def _extract_education_level(self, education: List[Any]) -> str:
        """Extract highest education level."""
        if not education:
            return 'Unknown'
        
        levels = []
        for edu in education:
            if isinstance(edu, dict):
                degree = edu.get('degree', '').lower()
                if 'phd' in degree or 'doctorate' in degree:
                    levels.append('PhD')
                elif 'master' in degree or 'mba' in degree:
                    levels.append('Masters')
                elif 'bachelor' in degree:
                    levels.append('Bachelors')
                elif degree:
                    levels.append('Other')
        
        if 'PhD' in levels:
            return 'PhD'
        elif 'Masters' in levels:
            return 'Masters'
        elif 'Bachelors' in levels:
            return 'Bachelors'
        else:
            return 'Other/Unknown'
    
    def _extract_schools(self, education: List[Any]) -> str:
        """Extract schools."""
        schools = []
        for edu in education:
            if isinstance(edu, dict):
                school = edu.get('school', '')
                if school:
                    schools.append(school)
        return ', '.join(schools)
    
    def _extract_degrees(self, education: List[Any]) -> str:
        """Extract degrees."""
        degrees = []
        for edu in education:
            if isinstance(edu, dict):
                degree = edu.get('degree', '')
                if degree:
                    degrees.append(degree)
        return ', '.join(degrees)
    
    def _categorize_experience_level(self, candidate: Dict[str, Any]) -> str:
        """Categorize experience level."""
        headline = candidate.get('headline', '').lower()
        
        if any(term in headline for term in ['senior', 'lead', 'principal', 'staff']):
            return 'Senior'
        elif any(term in headline for term in ['director', 'vp', 'head of', 'chief']):
            return 'Executive'
        elif any(term in headline for term in ['junior', 'associate', 'entry']):
            return 'Junior'
        else:
            return 'Mid-Level'
